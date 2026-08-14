import io
import csv
from fpdf import FPDF
from openpyxl import Workbook
from app.modules.election.schemas.results import ElectionResultSchema

class ExportService:
    @staticmethod
    def generate_csv(results: ElectionResultSchema) -> io.StringIO:
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["Election ID", str(results.election_id)])
        writer.writerow(["Generated At", results.generated_at.isoformat()])
        writer.writerow([])
        
        if results.statistics:
            writer.writerow(["Total Votes Cast", results.statistics.total_votes_cast])
            writer.writerow([])
        
        writer.writerow(["Category", "Candidate", "Votes", "Percentage", "Rank", "Is Winner", "Is Tied"])
        
        for category in (results.categories or []):
            for candidate in category.candidates:
                writer.writerow([
                    category.name,
                    candidate.name,
                    candidate.vote_count,
                    f"{candidate.percentage:.2f}%",
                    candidate.rank,
                    "Yes" if candidate.is_winner else "No",
                    "Yes" if candidate.is_tied else "No"
                ])
                
        output.seek(0)
        return output

    @staticmethod
    def generate_excel(results: ElectionResultSchema) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Election Results"
        
        ws.append(["Election ID", str(results.election_id)])
        ws.append(["Generated At", results.generated_at.isoformat()])
        ws.append([])
        
        ws.append(["Category", "Candidate", "Votes", "Percentage", "Rank", "Is Winner", "Is Tied"])
        
        for category in (results.categories or []):
            for candidate in category.candidates:
                ws.append([
                    category.name,
                    candidate.name,
                    candidate.vote_count,
                    f"{candidate.percentage:.2f}%",
                    candidate.rank,
                    "Yes" if candidate.is_winner else "No",
                    "Yes" if candidate.is_tied else "No"
                ])
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(results: ElectionResultSchema) -> io.BytesIO:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("helvetica", size=12)
        
        pdf.cell(200, 10, text="Election Results", align='C')
        pdf.ln(10)
        pdf.cell(200, 10, text=f"Election ID: {results.election_id}")
        pdf.ln(10)
        pdf.cell(200, 10, text=f"Generated At: {results.generated_at.isoformat()}")
        pdf.ln(10)
        
        for category in (results.categories or []):
            pdf.set_font("helvetica", style="B", size=11)
            pdf.cell(200, 10, text=f"Category: {category.name}")
            pdf.ln(10)
            pdf.set_font("helvetica", size=10)
            
            # Header
            pdf.cell(50, 10, "Candidate", 1)
            pdf.cell(30, 10, "Votes", 1)
            pdf.cell(30, 10, "Percentage", 1)
            pdf.cell(20, 10, "Rank", 1)
            pdf.cell(30, 10, "Winner?", 1)
            pdf.ln()
            
            for candidate in category.candidates:
                pdf.cell(50, 10, candidate.name, 1)
                pdf.cell(30, 10, str(candidate.vote_count), 1)
                pdf.cell(30, 10, f"{candidate.percentage:.2f}%", 1)
                pdf.cell(20, 10, str(candidate.rank), 1)
                pdf.cell(30, 10, "Yes" if candidate.is_winner else "No", 1)
                pdf.ln()
                
            pdf.ln(5)
            
        pdf_bytes = pdf.output(dest='S')
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin1')
            
        output = io.BytesIO(pdf_bytes)
        output.seek(0)
        return output
